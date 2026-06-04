"""PDF + Excel renderers for the unified report builder.

Both renderers share the BuzUp / TPM-TUR / UpDigital branding used in the
agent revenue admin exports (see `apps.agent_api.exporters`). We deliberately
mirror the look and feel so the operator gets consistent reports across all
admin areas.
"""
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

NAVY = colors.HexColor("#071E49")
ORANGE = colors.HexColor("#E47B11")
GREY = colors.HexColor("#6B6356")
LIGHT_GREY = colors.HexColor("#E7E1D4")
SOFT_BG = colors.HexColor("#F7F4EE")


def _asset(*parts):
    return Path(settings.BASE_DIR) / "static" / "assets" / Path(*parts)


def _safe_image(path: Path):
    try:
        if path.exists():
            return ImageReader(str(path))
    except Exception:
        return None
    return None


def _stringify(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if value is None:
        return ""
    return str(value)


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def render_pdf(
    title: str,
    period_from: str,
    period_to: str,
    columns: list[tuple[str, str]],
    rows: list[dict],
    totals: dict | None = None,
    filters_summary: str = "",
) -> bytes:
    buf = io.BytesIO()
    page = landscape(A4)
    width, height = page
    c = canvas.Canvas(buf, pagesize=page)

    _draw_header(c, width, height, title=title, period_from=period_from, period_to=period_to)

    y = height - 32 * mm

    if filters_summary:
        c.setFillColor(GREY)
        c.setFont("Helvetica", 9)
        c.drawString(10 * mm, y, f"Filtros: {filters_summary}")
        y -= 5 * mm

    if totals:
        y = _draw_totals(c, width, y, totals)

    # Compute column widths proportionally so all fit in landscape A4.
    total_chars = sum(max(len(label), 6) for _, label in columns) or 1
    avail = width - 20 * mm
    col_widths = [avail * (max(len(label), 6) / total_chars) for _, label in columns]

    y = _draw_table(c, x_left=10 * mm, y=y, width=width - 20 * mm,
                   columns=columns, col_widths=col_widths,
                   rows=rows, page_size=page)

    _draw_footer(c, width)
    c.save()
    return buf.getvalue()


def _draw_header(c, width, height, *, title, period_from, period_to):
    band_h = 22 * mm
    c.setFillColor(NAVY)
    c.rect(0, height - band_h, width, band_h, fill=1, stroke=0)

    logo = _safe_image(_asset("tpm-tur-logo", "tpm_dark.png")) or _safe_image(_asset("tpm-tur-logo", "tpm_light.png"))
    if logo:
        try:
            iw, ih = logo.getSize()
            target_h = 14 * mm
            target_w = iw * target_h / ih
            c.drawImage(logo, 8 * mm, height - band_h + (band_h - target_h) / 2,
                       width=target_w, height=target_h, mask="auto")
        except Exception:
            pass

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(60 * mm, height - 10 * mm, "BuzUp | Relatorio Administrativo")
    c.setFont("Helvetica", 9)
    c.drawString(60 * mm, height - 15 * mm, "TPM-TUR S.A. | Transporte cashless de Mocambique")

    c.setFillColor(ORANGE)
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(width - 10 * mm, height - 10 * mm, title)
    c.setFillColor(colors.white)
    c.setFont("Helvetica", 8)
    c.drawRightString(width - 10 * mm, height - 15 * mm,
                     f"Periodo: {period_from} - {period_to} | Gerado {datetime.now().strftime('%Y-%m-%d %H:%M')}")


def _draw_totals(c, width, y, totals):
    if not totals:
        return y
    box_w = (width - 20 * mm - (len(totals) - 1) * 4) / max(1, len(totals))
    box_h = 18 * mm
    x = 10 * mm
    for label, value in totals.items():
        c.setStrokeColor(LIGHT_GREY)
        c.setFillColor(colors.white)
        c.roundRect(x, y - box_h, box_w, box_h, 4, fill=1, stroke=1)
        c.setFillColor(GREY)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(x + 4, y - 6, label.replace("_", " ").upper())
        c.setFillColor(NAVY)
        c.setFont("Helvetica-Bold", 13)
        c.drawString(x + 4, y - 14, str(value))
        x += box_w + 4
    return y - box_h - 6 * mm


def _draw_table(c, *, x_left, y, width, columns, col_widths, rows, page_size):
    line_h = 5.5 * mm
    page_w, page_h = page_size

    c.setFillColor(NAVY)
    c.rect(x_left, y - line_h, width, line_h, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    x = x_left
    for (_, label), w in zip(columns, col_widths):
        c.drawString(x + 2, y - line_h + 1.5 * mm, label[:18])
        x += w
    y -= line_h

    if not rows:
        c.setFillColor(GREY)
        c.setFont("Helvetica-Oblique", 9)
        c.drawString(x_left + 4, y - 4 * mm, "Sem registos no periodo / filtros indicados.")
        return y - 8 * mm

    c.setFont("Helvetica", 7.5)
    for i, row in enumerate(rows):
        # Page break
        if y < 20 * mm:
            _draw_footer(c, page_w)
            c.showPage()
            _draw_header(c, page_w, page_h, title="(continuacao)", period_from="", period_to="")
            y = page_h - 30 * mm
            # Reprint header
            c.setFillColor(NAVY)
            c.rect(x_left, y - line_h, width, line_h, fill=1, stroke=0)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 8)
            x = x_left
            for (_, label), w in zip(columns, col_widths):
                c.drawString(x + 2, y - line_h + 1.5 * mm, label[:18])
                x += w
            y -= line_h
            c.setFont("Helvetica", 7.5)

        if i % 2 == 1:
            c.setFillColor(SOFT_BG)
            c.rect(x_left, y - line_h, width, line_h, fill=1, stroke=0)
        c.setFillColor(NAVY)
        x = x_left
        for (key, _), w in zip(columns, col_widths):
            txt = _stringify(row.get(key))
            # Truncate to fit the column visually
            max_chars = max(6, int(w / 2.2))
            if len(txt) > max_chars:
                txt = txt[: max_chars - 1] + "…"
            c.drawString(x + 2, y - line_h + 1.5 * mm, txt)
            x += w
        y -= line_h
    return y - 4 * mm


def _draw_footer(c, width):
    band_h = 14 * mm
    c.setFillColor(SOFT_BG)
    c.rect(0, 0, width, band_h, fill=1, stroke=0)
    up = _safe_image(_asset("up-digital-logo", "up_digital_dark.png")) or _safe_image(_asset("up-digital-logo", "up_digital_light.png"))
    if up:
        try:
            iw, ih = up.getSize()
            target_h = 7 * mm
            target_w = iw * target_h / ih
            c.drawImage(up, width - 8 * mm - target_w, (band_h - target_h) / 2,
                       width=target_w, height=target_h, mask="auto")
        except Exception:
            pass
    c.setFillColor(GREY)
    c.setFont("Helvetica", 8)
    c.drawString(10 * mm, band_h / 2 - 2, "BuzUp | TPM-TUR S.A. | Documento gerado automaticamente.")
    c.drawString(10 * mm, band_h / 2 - 9, "powered by")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def render_xlsx(
    title: str,
    period_from: str,
    period_to: str,
    columns: list[tuple[str, str]],
    rows: list[dict],
    totals: dict | None = None,
    filters_summary: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True, color="071E49")
    ws["A2"] = f"Periodo: {period_from} a {period_to}"
    ws["A3"] = f"Filtros: {filters_summary}" if filters_summary else ""
    ws["A4"] = f"Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    if totals:
        ws["A6"] = "Totais"
        ws["A6"].font = Font(bold=True)
        for i, (label, value) in enumerate(totals.items(), start=7):
            ws.cell(row=i, column=1, value=label.replace("_", " ").upper()).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 22

    # Data sheet
    data_ws = wb.create_sheet(title="Dados")
    head_fill = PatternFill("solid", fgColor="071E49")
    head_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style="thin", color="DDDDDD"),
                    right=Side(style="thin", color="DDDDDD"),
                    top=Side(style="thin", color="DDDDDD"),
                    bottom=Side(style="thin", color="DDDDDD"))

    for col_i, (_, label) in enumerate(columns, start=1):
        cell = data_ws.cell(row=1, column=col_i, value=label)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        data_ws.column_dimensions[get_column_letter(col_i)].width = max(14, min(40, len(label) + 8))

    for r_i, row in enumerate(rows, start=2):
        for c_i, (key, _) in enumerate(columns, start=1):
            value = row.get(key)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M")
            elif isinstance(value, (Decimal,)):
                value = float(value)
            cell = data_ws.cell(row=r_i, column=c_i, value=value)
            cell.border = border
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
    data_ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
