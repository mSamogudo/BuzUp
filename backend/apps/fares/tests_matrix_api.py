"""Endpoints da tabela de precos: quem pode mexer e o que grava.

Mudar precos e mudar quanto se cobra a toda a gente que anda naquela rota.
Por isso quem so consulta tarifas nao pode gravar a tabela, e uma importacao
nunca escreve sem passar pela pre-visualizacao.
"""

import io

from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework.test import APITestCase

from apps.fares import matrix as tabela
from apps.routes.models import Route, RouteStop, Stop
from apps.users.models import Role, UserRole


def _utilizador(username, permissoes):
    User = get_user_model()
    u = User.objects.create_user(
        username=username, email=f"{username}@exemplo.co.mz", password="x",
    )
    papel = Role.objects.create(name=username, code=username, permissions=permissoes)
    UserRole.objects.create(user=u, role=papel)
    return u


class MatrizApiBase(APITestCase):
    def setUp(self):
        self.rota = Route.objects.create(
            code="R-API", name="Maputo - Nelspruit",
            service_type=Route.ServiceType.INTERNATIONAL, status=Route.Status.ACTIVE,
        )
        self.paragens = []
        for i, (cod, nome) in enumerate(
            [("A-MAP", "Maputo"), ("A-MOA", "Moamba"), ("A-NEL", "Nelspruit")], start=1,
        ):
            s = Stop.objects.create(code=cod, name=nome, status="active")
            RouteStop.objects.create(route=self.rota, stop=s, sequence=i, direction="outbound")
            self.paragens.append(s)
        self.maputo, self.moamba, self.nelspruit = self.paragens

        self.gestor = _utilizador("gestor", ["fares.read", "fares.manage", "routes.manage"])
        self.leitor = _utilizador("leitor", ["fares.read"])

    def url(self, sufixo=""):
        return f"/api/admin/routes/{self.rota.id}/fare-matrix/{sufixo}"

    def _par(self, a, b):
        return f"{a.id}-{b.id}"


class LeituraTests(MatrizApiBase):
    def test_leitor_ve_a_tabela(self):
        self.client.force_authenticate(self.leitor)
        r = self.client.get(self.url())
        self.assertEqual(r.status_code, 200)
        self.assertEqual(len(r.data["stops"]), 3)
        self.assertEqual(r.data["pairs_total"], 6)

    def test_conta_os_trajectos_que_nao_se_vendem(self):
        self.client.force_authenticate(self.leitor)
        r = self.client.get(self.url())
        # So a ida e um trajecto valido enquanto nao houver sentido de volta,
        # e nenhum tem preco: 3 pares de ida sem tarifa.
        self.assertEqual(r.data["unsellable"], 3)
        self.assertFalse(r.data["has_return"])

    def test_rota_inexistente_da_404(self):
        self.client.force_authenticate(self.leitor)
        self.assertEqual(self.client.get("/api/admin/routes/999999/fare-matrix/").status_code, 404)

    def test_sem_sessao_e_recusado(self):
        self.assertIn(self.client.get(self.url()).status_code, (401, 403))


class GravacaoTests(MatrizApiBase):
    def test_gestor_grava_a_tabela(self):
        self.client.force_authenticate(self.gestor)
        r = self.client.post(self.url(), {
            "prices": {self._par(self.maputo, self.nelspruit): "2500"},
            "fallback_amount": "1000",
        }, format="json")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data["saved"]["created"], 1)
        self.assertEqual(r.data["fallback_amount"], "1000.00")
        self.assertEqual(r.data["unsellable"], 0, "com preco de recurso nada fica por vender")

    def test_quem_so_le_tarifas_nao_grava_precos(self):
        self.client.force_authenticate(self.leitor)
        r = self.client.post(self.url(), {
            "prices": {self._par(self.maputo, self.nelspruit): "2500"},
        }, format="json")
        self.assertEqual(r.status_code, 403)
        self.assertEqual(tabela.read_matrix(self.rota)["prices"], {})

    def test_preco_invalido_nao_grava_nada(self):
        self.client.force_authenticate(self.gestor)
        r = self.client.post(self.url(), {
            "prices": {
                self._par(self.maputo, self.moamba): "300",
                self._par(self.moamba, self.nelspruit): "abc",
            },
        }, format="json")
        self.assertEqual(r.status_code, 400)
        self.assertEqual(tabela.read_matrix(self.rota)["prices"], {})


class SentidoDeVoltaTests(MatrizApiBase):
    def test_criar_volta_torna_o_regresso_um_trajecto(self):
        self.client.force_authenticate(self.gestor)
        r = self.client.post(self.url("return-direction/"), {}, format="json")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data["return_created"]["created"], 3)
        self.assertTrue(r.data["has_return"])
        self.assertEqual(r.data["unsellable"], 6, "agora sao 6 trajectos, todos sem preco")

    def test_o_regresso_passa_a_resolver_logo_a_seguir(self):
        """A cache de segmentos guarda tambem os pares que NAO existem.

        Sem a limpar, o regresso continuava recusado depois de criado o
        sentido de volta — e o botao parecia nao ter feito nada.
        """
        from apps.routes.services import resolve_route_segment

        self.client.force_authenticate(self.gestor)
        # Aquece a cache com o "nao existe" de antes da volta existir.
        with self.assertRaises(Exception):
            resolve_route_segment(self.rota, self.nelspruit.id, self.maputo.id)

        self.client.post(self.url("return-direction/"), {}, format="json")

        self.assertIsNotNone(resolve_route_segment(self.rota, self.nelspruit.id, self.maputo.id))

    def test_repetir_nao_duplica_paragens(self):
        self.client.force_authenticate(self.gestor)
        self.client.post(self.url("return-direction/"), {}, format="json")
        r = self.client.post(self.url("return-direction/"), {}, format="json")
        self.assertTrue(r.data["return_created"]["already"])
        self.assertEqual(RouteStop.objects.filter(route=self.rota, direction="inbound").count(), 3)


class PreenchimentoTests(MatrizApiBase):
    def test_preencher_devolve_sugestao_sem_gravar(self):
        self.client.force_authenticate(self.gestor)
        r = self.client.post(self.url("fill/"), {"base": "100", "per_stop": "50"}, format="json")
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.data["prices"])
        self.assertEqual(
            tabela.read_matrix(self.rota)["prices"], {},
            "uma sugestao aplicada em silencio e impossivel de rever depois",
        )


class ModeloEImportacaoTests(MatrizApiBase):
    def test_modelo_par_a_par_traz_uma_linha_por_trajecto(self):
        self.client.force_authenticate(self.leitor)
        r = self.client.get(self.url("template/"), {"method": "origin_destination"})
        self.assertEqual(r.status_code, 200)
        ws = load_workbook(io.BytesIO(r.content)).active
        linhas = [l for l in range(5, ws.max_row + 1) if ws.cell(row=l, column=1).value]
        self.assertEqual(len(linhas), 6)

    def test_modelo_de_preco_unico_traz_uma_linha_so(self):
        self.client.force_authenticate(self.leitor)
        r = self.client.get(self.url("template/"), {"method": "fixed"})
        ws = load_workbook(io.BytesIO(r.content)).active
        self.assertEqual(ws.cell(row=4, column=1).value, "aplica_a")
        self.assertIsNone(ws.cell(row=6, column=1).value)

    def test_metodo_desconhecido_da_400(self):
        self.client.force_authenticate(self.leitor)
        self.assertEqual(self.client.get(self.url("template/"), {"method": "por_km"}).status_code, 400)

    def _folha_preenchida(self, method="origin_destination", valor=750):
        wb = load_workbook(io.BytesIO(tabela.template_xlsx(self.rota, method=method)))
        ws = wb.active
        if method == "fixed":
            ws.cell(row=5, column=2, value=valor)
        else:
            for l in range(5, ws.max_row + 1):
                if ws.cell(row=l, column=1).value:
                    ws.cell(row=l, column=5, value=valor)
        buf = io.BytesIO()
        wb.save(buf)
        return io.BytesIO(buf.getvalue())

    def _enviar(self, folha, **extra):
        folha.name = "precos.xlsx"
        return self.client.post(
            self.url("import/"), {"file": folha, **extra}, format="multipart",
        )

    def test_importacao_pre_visualiza_sem_gravar(self):
        self.client.force_authenticate(self.gestor)
        r = self._enviar(self._folha_preenchida())
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.data["preview"])
        self.assertEqual(r.data["changes"], 6)
        self.assertEqual(
            tabela.read_matrix(self.rota)["prices"], {},
            "uma tabela aplicada as cegas e dinheiro cobrado a mais ou a menos",
        )

    def test_importacao_com_apply_grava(self):
        self.client.force_authenticate(self.gestor)
        r = self._enviar(self._folha_preenchida(), apply="true")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data["saved"]["created"], 6)
        self.assertEqual(len(tabela.read_matrix(self.rota)["prices"]), 6)

    def test_importar_preco_unico_define_o_recurso(self):
        self.client.force_authenticate(self.gestor)
        r = self._enviar(self._folha_preenchida(method="fixed", valor=900), apply="true")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data["fallback_amount"], "900.00")
        self.assertEqual(r.data["prices"], {})

    def test_ficheiro_que_nao_e_o_modelo_e_recusado(self):
        self.client.force_authenticate(self.gestor)
        falso = io.BytesIO(b"isto nao e um excel")
        falso.name = "precos.xlsx"
        r = self._enviar(falso)
        self.assertEqual(r.status_code, 400)

    def test_sem_ficheiro_da_400(self):
        self.client.force_authenticate(self.gestor)
        self.assertEqual(self.client.post(self.url("import/"), {}, format="multipart").status_code, 400)

    def test_quem_so_le_nao_importa(self):
        self.client.force_authenticate(self.leitor)
        r = self._enviar(self._folha_preenchida(), apply="true")
        self.assertEqual(r.status_code, 403)
