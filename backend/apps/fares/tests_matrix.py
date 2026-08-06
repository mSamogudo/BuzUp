"""Tabela de precos de uma rota: grelha, Excel e o que nao se pode vender.

Porque isto importa: a MZ-NEL, a rota que o piloto vai usar, tinha UM par de
paragens com preco. Os outros 14 trajectos falhavam com "Nenhuma tarifa
configurada" — o passageiro escolhia origem e destino e nao conseguia comprar.
O motor sempre soube cobrar por par; o que faltava era conseguir configurar
812 pares sem ser um a um.

Um erro nesta tabela e dinheiro cobrado a mais ou a menos em todas as viagens
ate alguem reparar. Por isso: transaccao unica ao gravar, recusa da folha
inteira ao primeiro problema, e pre-visualizacao antes de aplicar.
"""

from decimal import Decimal

from django.test import TestCase

from apps.fares import matrix as tabela
from apps.fares.models import FareProduct, FareRule
from apps.fares.services import NoFareFoundError, quote_fare
from apps.routes.models import Route, RouteStop, Stop


class TabelaBase(TestCase):
    def setUp(self):
        self.rota = Route.objects.create(
            code="R-TAB", name="Maputo - Nelspruit",
            service_type=Route.ServiceType.INTERNATIONAL, status=Route.Status.ACTIVE,
        )
        self.paragens = []
        for i, (cod, nome) in enumerate(
            [("P-MAP", "Maputo"), ("P-MOA", "Moamba"), ("P-RES", "Ressano"), ("P-NEL", "Nelspruit")],
            start=1,
        ):
            s = Stop.objects.create(code=cod, name=nome, status="active")
            RouteStop.objects.create(route=self.rota, stop=s, sequence=i, direction="outbound")
            self.paragens.append(s)
        self.maputo, self.moamba, self.ressano, self.nelspruit = self.paragens
        # Sem sentido de volta o regresso nem e um trajecto valido — ver
        # `ensure_return_direction`.
        tabela.ensure_return_direction(self.rota)

    def _par(self, a, b):
        return f"{a.id}-{b.id}"


class LerEGravarTests(TabelaBase):
    def test_tabela_nova_esta_vazia_e_conta_os_pares(self):
        m = tabela.read_matrix(self.rota)
        self.assertEqual(m["prices"], {})
        # 4 paragens = 12 pares com sentido (ida e volta contam separado).
        self.assertEqual(m["pairs_total"], 12)
        self.assertEqual(m["pairs_priced"], 0)

    def test_gravar_um_par_torna_o_trajecto_vendavel(self):
        with self.assertRaises(NoFareFoundError):
            quote_fare(route=self.rota, origin_stop=self.maputo, destination_stop=self.nelspruit)

        tabela.write_matrix(self.rota, {self._par(self.maputo, self.nelspruit): "2500"})

        q = quote_fare(route=self.rota, origin_stop=self.maputo, destination_stop=self.nelspruit)
        self.assertEqual(q.amount, Decimal("2500.00"))

    def test_ida_e_volta_sao_precos_independentes(self):
        """`A->B` e `B->A` sao regras distintas — e assim que se cobra
        diferente ao regresso, se o operador quiser."""
        tabela.write_matrix(self.rota, {
            self._par(self.maputo, self.nelspruit): "2500",
            self._par(self.nelspruit, self.maputo): "2300",
        })
        ida = quote_fare(route=self.rota, origin_stop=self.maputo, destination_stop=self.nelspruit)
        volta = quote_fare(route=self.rota, origin_stop=self.nelspruit, destination_stop=self.maputo)
        self.assertEqual(ida.amount, Decimal("2500.00"))
        self.assertEqual(volta.amount, Decimal("2300.00"))

    def test_gravar_de_novo_actualiza_em_vez_de_duplicar(self):
        par = self._par(self.maputo, self.moamba)
        tabela.write_matrix(self.rota, {par: "300"})
        r = tabela.write_matrix(self.rota, {par: "350"})
        self.assertEqual(r["updated"], 1)
        self.assertEqual(
            FareRule.objects.filter(
                route=self.rota, origin_stop=self.maputo, destination_stop=self.moamba,
            ).count(), 1,
            "duas regras para o mesmo par dariam conflito de tarifas na compra",
        )

    def test_valor_vazio_apaga_o_preco_do_par(self):
        par = self._par(self.maputo, self.moamba)
        tabela.write_matrix(self.rota, {par: "300"})
        r = tabela.write_matrix(self.rota, {par: ""})
        self.assertEqual(r["deleted"], 1)
        self.assertEqual(tabela.read_matrix(self.rota)["prices"], {})

    def test_paragem_de_outra_rota_e_recusada(self):
        intrusa = Stop.objects.create(code="P-XXX", name="Fora", status="active")
        with self.assertRaises(tabela.MatrixError) as ctx:
            tabela.write_matrix(self.rota, {f"{self.maputo.id}-{intrusa.id}": "100"})
        self.assertIn("nao pertence", str(ctx.exception))

    def test_preco_negativo_e_recusado(self):
        with self.assertRaises(tabela.MatrixError):
            tabela.write_matrix(self.rota, {self._par(self.maputo, self.moamba): "-50"})

    def test_uma_linha_ma_nao_deixa_meia_tabela_gravada(self):
        """Meia tabela aplicada e pior do que nenhuma: ninguem sabe que metade
        ficou por aplicar."""
        with self.assertRaises(tabela.MatrixError):
            tabela.write_matrix(self.rota, {
                self._par(self.maputo, self.moamba): "300",
                self._par(self.moamba, self.ressano): "abc",
            })
        self.assertEqual(tabela.read_matrix(self.rota)["prices"], {})


class PrecoDeRecursoTests(TabelaBase):
    def test_recurso_torna_todos_os_trajectos_vendaveis(self):
        # E a rede de seguranca que faltava a MZ-NEL: sem ela, qualquer par
        # sem preco proprio simplesmente nao se vende.
        tabela.write_matrix(self.rota, {}, fallback_amount="1000")
        q = quote_fare(route=self.rota, origin_stop=self.moamba, destination_stop=self.ressano)
        self.assertEqual(q.amount, Decimal("1000.00"))

    def test_preco_do_par_manda_sobre_o_recurso(self):
        tabela.write_matrix(
            self.rota, {self._par(self.maputo, self.nelspruit): "2500"}, fallback_amount="1000",
        )
        especifico = quote_fare(route=self.rota, origin_stop=self.maputo, destination_stop=self.nelspruit)
        generico = quote_fare(route=self.rota, origin_stop=self.moamba, destination_stop=self.ressano)
        self.assertEqual(especifico.amount, Decimal("2500.00"))
        self.assertEqual(generico.amount, Decimal("1000.00"))


class PreenchimentoPorParagensTests(TabelaBase):
    def test_preco_cresce_com_o_numero_de_paragens(self):
        precos = tabela.fill_by_distance(self.rota, base="100", per_stop="50")
        # Maputo -> Moamba e um salto; Maputo -> Nelspruit sao tres.
        self.assertEqual(precos[self._par(self.maputo, self.moamba)], "100.00")
        self.assertEqual(precos[self._par(self.maputo, self.nelspruit)], "200.00")

    def test_e_simetrico_por_omissao(self):
        precos = tabela.fill_by_distance(self.rota, base="100", per_stop="50")
        self.assertEqual(
            precos[self._par(self.maputo, self.nelspruit)],
            precos[self._par(self.nelspruit, self.maputo)],
        )

    def test_cobre_todos_os_pares(self):
        precos = tabela.fill_by_distance(self.rota, base="100", per_stop="50")
        self.assertEqual(len(precos), 12)


class ExcelTests(TabelaBase):
    def test_modelo_traz_uma_linha_por_par_com_os_codigos_preenchidos(self):
        from openpyxl import load_workbook
        import io

        tabela.write_matrix(self.rota, {self._par(self.maputo, self.nelspruit): "2500"})
        wb = load_workbook(io.BytesIO(tabela.template_xlsx(self.rota)))
        ws = wb.active

        linhas = [
            [ws.cell(row=r, column=c).value for c in range(1, 6)]
            for r in range(5, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        self.assertEqual(len(linhas), 12)
        # O preco que ja existe vem preenchido, para nao se perder ao reimportar.
        ja = [l for l in linhas if l[0] == "P-MAP" and l[2] == "P-NEL"]
        self.assertEqual(ja[0][4], "2500.00")

    def test_importar_o_modelo_preenchido_devolve_os_pares(self):
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(tabela.template_xlsx(self.rota)))
        ws = wb.active
        for r in range(5, ws.max_row + 1):
            if ws.cell(row=r, column=1).value:
                ws.cell(row=r, column=5, value=750)
        buf = io.BytesIO()
        wb.save(buf)

        lido = tabela.parse_xlsx(self.rota, buf.getvalue())
        self.assertEqual(len(lido["prices"]), 12)
        tabela.write_matrix(self.rota, lido["prices"])
        q = quote_fare(route=self.rota, origin_stop=self.moamba, destination_stop=self.nelspruit)
        self.assertEqual(q.amount, Decimal("750.00"))

    def test_codigo_de_paragem_estranha_recusa_a_folha_inteira(self):
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(tabela.template_xlsx(self.rota)))
        ws = wb.active
        ws.cell(row=5, column=1, value="P-INVENTADA")
        buf = io.BytesIO()
        wb.save(buf)

        with self.assertRaises(tabela.MatrixError) as ctx:
            tabela.parse_xlsx(self.rota, buf.getvalue())
        self.assertIn("nao pertence", str(ctx.exception))

    def test_ficheiro_que_nao_e_o_modelo_e_recusado_com_instrucao(self):
        with self.assertRaises(tabela.MatrixError) as ctx:
            tabela.parse_xlsx(self.rota, b"isto nao e um excel")
        self.assertIn("modelo", str(ctx.exception).lower())

    def test_preco_com_virgula_decimal_e_aceite(self):
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(tabela.template_xlsx(self.rota)))
        ws = wb.active
        ws.cell(row=5, column=5, value="1.250,50")
        buf = io.BytesIO()
        wb.save(buf)
        # Nao rebenta a ler; o valor e normalizado na gravacao.
        lido = tabela.parse_xlsx(self.rota, buf.getvalue())
        self.assertTrue(lido["prices"])


class ModeloPrecoUnicoTests(TabelaBase):
    """O modelo do metodo `fixed`: uma linha, um preco, a rota inteira."""

    def test_modelo_fixo_tem_uma_so_linha_de_preco(self):
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(tabela.template_xlsx(self.rota, method="fixed")))
        ws = wb.active
        self.assertEqual(ws.cell(row=4, column=1).value, "aplica_a")
        self.assertIsNone(ws.cell(row=6, column=1).value, "o modelo fixo nao lista pares")

    def test_importar_o_modelo_fixo_da_preco_a_todos_os_trajectos(self):
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(tabela.template_xlsx(self.rota, method="fixed")))
        ws = wb.active
        ws.cell(row=5, column=2, value=900)
        buf = io.BytesIO()
        wb.save(buf)

        lido = tabela.parse_xlsx(self.rota, buf.getvalue())
        self.assertEqual(lido["prices"], {})
        tabela.write_matrix(
            self.rota, lido["prices"], fallback_amount=lido["fallback_amount"],
        )
        q = quote_fare(route=self.rota, origin_stop=self.moamba, destination_stop=self.nelspruit)
        self.assertEqual(q.amount, Decimal("900.00"))

    def test_modelo_fixo_por_preencher_e_recusado(self):
        with self.assertRaises(tabela.MatrixError):
            tabela.parse_xlsx(self.rota, tabela.template_xlsx(self.rota, method="fixed"))

    def test_metodo_desconhecido_e_recusado(self):
        with self.assertRaises(tabela.MatrixError):
            tabela.template_xlsx(self.rota, method="por_km")
