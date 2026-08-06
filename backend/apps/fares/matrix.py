"""Tabela de precos de uma rota: ler, gravar, exportar e importar.

**Porque existe.** O motor de tarifas ja sabia cobrar por par de paragens
(`FareRule` com `ORIGIN_DESTINATION`), mas so se podia criar uma regra de cada
vez. Numa rota com 29 paragens sao 812 pares — ninguem preenche isso a mao, e
por isso na pratica estava tudo a preco fixo. A MZ-NEL, que e a rota que o
piloto vai usar, tinha UM par com preco: os outros 14 trajectos nao se podiam
vender.

Este modulo trata a tabela como aquilo que ela e para quem a gere — uma grelha
origem x destino — e traduz para as regras que o motor ja usa. Nao inventa um
motor novo.

**Ida e volta.** `A->B` e `B->A` sao regras distintas no motor, portanto podem
ter precos diferentes. Por omissao a grelha e simetrica (o mesmo preco nos dois
sentidos), porque e o caso normal; quem quiser cobrar diferente ao regresso
edita a celula do sentido inverso.
"""

from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation

from django.db import transaction

from apps.fares.models import FareProduct, FareRule
from apps.routes.models import Route, RouteStop, Stop


class MatrixError(ValueError):
    """Problema na tabela que o utilizador tem de corrigir."""


def _to_decimal(valor) -> Decimal | None:
    if valor is None:
        return None
    texto = str(valor).strip().replace(" ", "").replace("\u00a0", "")
    # "1.250,50" e como se escreve mil duzentos e cinquenta e meio em portugues:
    # o ponto separa milhares e a virgula os decimais. Trocar so a virgula por
    # ponto dava "1.250.50", que nao e numero nenhum — e o operador via o
    # ficheiro inteiro recusado por ter escrito o preco como sempre escreve.
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", ".")
    if not texto:
        return None
    try:
        d = Decimal(texto)
    except (InvalidOperation, ValueError):
        raise MatrixError(f"Valor invalido: {valor!r}. Use apenas numeros (ex.: 250 ou 250.00).")
    if d < 0:
        raise MatrixError(f"Preco negativo nao faz sentido: {valor!r}.")
    return d.quantize(Decimal("0.01"))


def route_stops(route: Route) -> list[Stop]:
    """As paragens da rota, sem repetir, pela ordem em que o autocarro passa.

    Uma paragem pode aparecer na ida e na volta; para a tabela de precos
    interessa a paragem, nao o sentido — o sentido esta no par (A->B ou B->A).
    """
    vistos: dict[int, Stop] = {}
    for rs in RouteStop.objects.filter(route=route).select_related("stop").order_by("direction", "sequence"):
        vistos.setdefault(rs.stop_id, rs.stop)
    return list(vistos.values())


def ensure_return_direction(route: Route) -> dict:
    """Cria o sentido de VOLTA espelhando a ida, quando ele nao existe.

    Sem paragens de volta registadas, o regresso nao e sequer um trajecto
    valido: `resolve_route_segment` recusa-o com "destino deve estar depois da
    origem", muito antes de se falar de precos. Era o caso da MZ-NEL — um
    autocarro internacional que so podia ser vendido num sentido.

    A volta e a ida ao contrario, que e o que acontece na estrada. Se o
    percurso de regresso for mesmo diferente, edita-se depois; o que nao pode
    e nao existir.
    """
    ida = list(
        RouteStop.objects.filter(route=route, direction=RouteStop.Direction.OUTBOUND).order_by("sequence")
    )
    if not ida:
        raise MatrixError(f"A rota {route.code} nao tem paragens de ida para espelhar.")
    if RouteStop.objects.filter(route=route, direction=RouteStop.Direction.INBOUND).exists():
        return {"created": 0, "already": True}

    total_km = ida[-1].distance_from_start_km or Decimal("0.00")
    criadas = 0
    for nova_seq, rs in enumerate(reversed(ida), start=1):
        RouteStop.objects.create(
            route=route, stop=rs.stop, sequence=nova_seq,
            direction=RouteStop.Direction.INBOUND,
            distance_from_start_km=(total_km - (rs.distance_from_start_km or Decimal("0.00"))),
        )
        criadas += 1

    # Sem isto a volta continua a ser recusada: o resolvedor guarda em cache o
    # "este par nao forma segmento" de quando o sentido ainda nao existia, e o
    # botao no portal parecia nao ter feito nada.
    from apps.routes.services import invalidate_route_segments

    invalidate_route_segments(route.id)
    return {"created": criadas, "already": False}


def _single_trip_product(route: Route) -> FareProduct:
    """O produto onde as regras desta rota vivem.

    Reutiliza o que a rota ja usa; so cria um quando nao ha nenhum, para nao
    espalhar produtos duplicados por cada gravacao da tabela.
    """
    existente = (
        FareRule.objects.filter(
            route=route, fare_product__product_type=FareProduct.ProductType.SINGLE_TRIP,
        )
        .select_related("fare_product")
        .first()
    )
    if existente:
        return existente.fare_product
    produto, _ = FareProduct.objects.get_or_create(
        name=f"Avulso {route.code}",
        product_type=FareProduct.ProductType.SINGLE_TRIP,
        defaults={"status": FareProduct.Status.ACTIVE},
    )
    return produto


def read_matrix(route: Route, passenger_class: str = FareRule.PassengerClass.STANDARD) -> dict:
    """A tabela como o portal a mostra: paragens, precos por par e recurso."""
    paragens = route_stops(route)
    regras = FareRule.objects.filter(
        route=route,
        calculation_method=FareRule.CalculationMethod.ORIGIN_DESTINATION,
        passenger_class=passenger_class,
        origin_stop__isnull=False, destination_stop__isnull=False,
        fare_product__product_type=FareProduct.ProductType.SINGLE_TRIP,
    )
    precos = {
        f"{r.origin_stop_id}-{r.destination_stop_id}": str(r.fixed_amount)
        for r in regras
    }

    recurso = FareRule.objects.filter(
        route=route,
        calculation_method=FareRule.CalculationMethod.FIXED,
        passenger_class=passenger_class,
        fare_product__product_type=FareProduct.ProductType.SINGLE_TRIP,
    ).first()

    return {
        "route": {"id": route.id, "code": route.code, "name": route.name,
                  "service_type": route.service_type},
        "passenger_class": passenger_class,
        "stops": [{"id": s.id, "code": s.code, "name": s.name} for s in paragens],
        "prices": precos,
        # Sem regra de recurso, um par sem preco NAO SE VENDE — foi o que
        # deixou 14 dos 15 trajectos da MZ-NEL por vender.
        "fallback_amount": str(recurso.fixed_amount) if recurso else "",
        "pairs_total": max(len(paragens) * (len(paragens) - 1), 0),
        "pairs_priced": len(precos),
        # Sem sentido de volta o regresso nem chega a ser um trajecto valido,
        # por muito preco que a tabela tenha. O portal precisa de o dizer.
        "has_return": RouteStop.objects.filter(
            route=route, direction=RouteStop.Direction.INBOUND,
        ).exists(),
    }


@transaction.atomic
def write_matrix(
    route: Route,
    prices: dict,
    *,
    fallback_amount=None,
    passenger_class: str = FareRule.PassengerClass.STANDARD,
) -> dict:
    """Grava a tabela. `prices` e {"<origem>-<destino>": valor}.

    Um par com valor vazio APAGA a regra desse par (volta a cair no recurso).
    E gravada a tabela inteira de uma vez, dentro de uma transaccao: meia
    tabela gravada seria pior do que nenhuma.
    """
    produto = _single_trip_product(route)
    validos = {s.id for s in route_stops(route)}

    criadas = actualizadas = apagadas = 0
    for chave, valor in (prices or {}).items():
        try:
            origem_id, destino_id = (int(x) for x in str(chave).split("-", 1))
        except (TypeError, ValueError):
            raise MatrixError(f"Par invalido na tabela: {chave!r}.")
        if origem_id == destino_id:
            continue
        if origem_id not in validos or destino_id not in validos:
            raise MatrixError(
                f"O par {chave} usa uma paragem que nao pertence a rota {route.code}."
            )

        montante = _to_decimal(valor)
        existente = FareRule.objects.filter(
            route=route, fare_product=produto,
            calculation_method=FareRule.CalculationMethod.ORIGIN_DESTINATION,
            origin_stop_id=origem_id, destination_stop_id=destino_id,
            passenger_class=passenger_class,
        ).first()

        if montante is None:
            if existente:
                existente.delete()
                apagadas += 1
            continue

        if existente:
            if existente.fixed_amount != montante:
                existente.fixed_amount = montante
                existente.save(update_fields=["fixed_amount", "updated_at"])
                actualizadas += 1
        else:
            FareRule.objects.create(
                fare_product=produto, route=route,
                calculation_method=FareRule.CalculationMethod.ORIGIN_DESTINATION,
                origin_stop_id=origem_id, destination_stop_id=destino_id,
                passenger_class=passenger_class, fixed_amount=montante,
            )
            criadas += 1

    if fallback_amount is not None:
        _write_fallback(route, produto, fallback_amount, passenger_class)

    return {"created": criadas, "updated": actualizadas, "deleted": apagadas}


def _write_fallback(route: Route, produto: FareProduct, valor, passenger_class: str) -> None:
    """Preco de recurso da rota: o que se cobra num par sem preco proprio.

    Existe para nenhum trajecto ficar invendavel. Vazio remove-o — mas ai um
    par sem preco volta a ser recusado na compra.
    """
    montante = _to_decimal(valor)
    existente = FareRule.objects.filter(
        route=route, fare_product=produto,
        calculation_method=FareRule.CalculationMethod.FIXED,
        passenger_class=passenger_class,
    ).first()
    if montante is None:
        if existente:
            existente.delete()
        return
    if existente:
        if existente.fixed_amount != montante:
            existente.fixed_amount = montante
            existente.save(update_fields=["fixed_amount", "updated_at"])
    else:
        FareRule.objects.create(
            fare_product=produto, route=route,
            calculation_method=FareRule.CalculationMethod.FIXED,
            passenger_class=passenger_class, fixed_amount=montante,
        )


def fill_by_distance(route: Route, base, per_stop) -> dict:
    """Preenche a grelha a partir do numero de paragens entre origem e destino.

    Nao substitui a decisao comercial — e um ponto de partida para depois se
    afinarem os pares que fogem a regra. Sem isto, uma rota de 29 paragens
    (812 pares) nunca sai do preco unico.
    """
    base_d = _to_decimal(base) or Decimal("0.00")
    passo = _to_decimal(per_stop) or Decimal("0.00")

    ordem: dict[int, int] = {}
    for rs in RouteStop.objects.filter(
        route=route, direction=RouteStop.Direction.OUTBOUND,
    ).order_by("sequence"):
        ordem.setdefault(rs.stop_id, rs.sequence)
    # Rotas so com sentido de volta registado nao podem ficar de fora.
    if not ordem:
        for rs in RouteStop.objects.filter(route=route).order_by("direction", "sequence"):
            ordem.setdefault(rs.stop_id, rs.sequence)

    precos: dict[str, str] = {}
    paragens = route_stops(route)
    for a in paragens:
        for b in paragens:
            if a.id == b.id:
                continue
            saltos = abs(ordem.get(a.id, 0) - ordem.get(b.id, 0)) or 1
            precos[f"{a.id}-{b.id}"] = str((base_d + passo * (saltos - 1)).quantize(Decimal("0.01")))
    return precos


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

_CABECALHO = ("origem_codigo", "origem_nome", "destino_codigo", "destino_nome", "preco_mzn")
# O modelo do preco unico nao tem pares: uma linha so, que vale para a rota
# inteira. Cabecalho diferente para o leitor saber logo qual dos dois recebeu.
_CABECALHO_FIXO = ("aplica_a", "preco_mzn")
_LINHA_FIXA = "todos os trajectos"

METHODS = ("origin_destination", "fixed")


def template_xlsx(route: Route, *, method: str = "origin_destination") -> bytes:
    """Modelo em Excel para o operador preencher fora do portal.

    Dois modelos, porque sao duas maneiras diferentes de cobrar:

    - `origin_destination`: uma linha por par de paragens, com os codigos ja
      preenchidos — o operador so escreve o preco. Escrever os pares a mao era
      o caminho garantido para trocar codigos e criar precos em paragens que
      nao pertencem a rota.
    - `fixed`: uma linha unica. O mesmo preco em qualquer trajecto da rota.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    if method not in METHODS:
        raise MatrixError(f"Metodo desconhecido: {method!r}.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Precos"

    def _cabecalho(titulos, linha):
        for i, titulo in enumerate(titulos, start=1):
            c = ws.cell(row=linha, column=i, value=titulo)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="071E49")
            c.alignment = Alignment(horizontal="center")

    ws["A1"] = f"Tabela de precos — {route.code} · {route.name}"
    ws["A1"].font = Font(bold=True, size=13)

    if method == "fixed":
        ws["A2"] = (
            "Preco unico: o mesmo valor em qualquer trajecto desta rota. "
            "Preencha so a celula do preco."
        )
        ws["A2"].font = Font(italic=True, size=9)
        ws.append([])
        _cabecalho(_CABECALHO_FIXO, 4)
        ws.cell(row=5, column=1, value=_LINHA_FIXA)
        ws.cell(row=5, column=2, value=read_matrix(route)["fallback_amount"])
        for col, largura in zip("AB", (28, 14)):
            ws.column_dimensions[col].width = largura
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    ws["A2"] = (
        "Preencha apenas a coluna 'preco_mzn'. Deixar em branco = sem preco proprio "
        "(usa o preco de recurso da rota). Nao mude os codigos nem a ordem das colunas."
    )
    ws["A2"].font = Font(italic=True, size=9)
    ws.append([])

    cabecalho_linha = 4
    _cabecalho(_CABECALHO, cabecalho_linha)

    existentes = read_matrix(route)["prices"]
    paragens = route_stops(route)
    linha = cabecalho_linha + 1
    for a in paragens:
        for b in paragens:
            if a.id == b.id:
                continue
            ws.cell(row=linha, column=1, value=a.code)
            ws.cell(row=linha, column=2, value=a.name)
            ws.cell(row=linha, column=3, value=b.code)
            ws.cell(row=linha, column=4, value=b.name)
            ws.cell(row=linha, column=5, value=existentes.get(f"{a.id}-{b.id}", ""))
            linha += 1

    for col, largura in zip("ABCDE", (18, 30, 18, 30, 14)):
        ws.column_dimensions[col].width = largura
    ws.freeze_panes = ws.cell(row=cabecalho_linha + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_xlsx(route: Route, conteudo: bytes) -> dict:
    """Le o Excel preenchido e devolve `{"prices": {...}, "fallback_amount": ...}`.

    Aceita os dois modelos: o da tabela par a par e o do preco unico. Recusa a
    folha inteira ao primeiro problema, em vez de importar metade: uma tabela
    de precos meio aplicada e pior do que uma por aplicar, porque ninguem sabe
    que metade ficou.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception:
        raise MatrixError("Nao foi possivel ler o ficheiro. Use o modelo em Excel (.xlsx).")
    ws = wb.active

    por_codigo = {s.code: s for s in route_stops(route)}

    cabecalho_linha = None
    modelo = None
    for linha in range(1, min(ws.max_row, 20) + 1):
        valores = [str(ws.cell(row=linha, column=c).value or "").strip().lower() for c in range(1, 6)]
        if valores[:4] == list(_CABECALHO[:4]):
            cabecalho_linha, modelo = linha, "origin_destination"
            break
        if valores[:2] == list(_CABECALHO_FIXO):
            cabecalho_linha, modelo = linha, "fixed"
            break
    if cabecalho_linha is None:
        raise MatrixError(
            "Cabecalho nao encontrado. Descarregue o modelo e preencha so a coluna do preco."
        )

    if modelo == "fixed":
        bruto = ws.cell(row=cabecalho_linha + 1, column=2).value
        if bruto is None or str(bruto).strip() == "":
            raise MatrixError("O ficheiro nao tem nenhum preco preenchido.")
        try:
            _to_decimal(bruto)
        except MatrixError as e:
            raise MatrixError(f"Linha {cabecalho_linha + 1}: {e}")
        return {"prices": {}, "fallback_amount": str(bruto)}

    precos: dict[str, str] = {}
    for linha in range(cabecalho_linha + 1, ws.max_row + 1):
        origem = str(ws.cell(row=linha, column=1).value or "").strip()
        destino = str(ws.cell(row=linha, column=3).value or "").strip()
        if not origem and not destino:
            continue
        if origem not in por_codigo:
            raise MatrixError(f"Linha {linha}: a paragem de origem {origem!r} nao pertence a rota {route.code}.")
        if destino not in por_codigo:
            raise MatrixError(f"Linha {linha}: a paragem de destino {destino!r} nao pertence a rota {route.code}.")
        if origem == destino:
            continue
        bruto = ws.cell(row=linha, column=5).value
        try:
            _to_decimal(bruto)
        except MatrixError as e:
            raise MatrixError(f"Linha {linha}: {e}")
        precos[f"{por_codigo[origem].id}-{por_codigo[destino].id}"] = (
            "" if bruto is None or str(bruto).strip() == "" else str(bruto)
        )

    if not precos:
        raise MatrixError("O ficheiro nao tem nenhuma linha de preco preenchida.")
    return {"prices": precos, "fallback_amount": None}
