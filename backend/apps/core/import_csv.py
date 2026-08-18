from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def parse_excel_upload(file_content: bytes, required_fields: list[str], header_map: dict | None = None) -> tuple[list[dict], list[dict]]:
    wb = load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active
    rows_data = []
    errors = []
    hmap = header_map or {}

    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            raw = [str(c or "").strip().lower() for c in row]
            headers = [hmap.get(h, h) for h in raw]
            continue
        cleaned = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                cleaned[headers[j]] = str(val or "").strip()
        missing = [f for f in required_fields if f not in cleaned or not cleaned[f]]
        if missing:
            errors.append({"row": i, "detail": f"Campos obrigatorios em falta: {', '.join(missing)}"})
            continue
        rows_data.append(cleaned)

    wb.close()
    return rows_data, errors


CARD_HEADER_MAP = {
    "uid do cartao (obrigatorio)": "card_uid",
    "uid do cartao": "card_uid",
    "card_uid": "card_uid",
    "lote": "issued_batch",
    "issued_batch": "issued_batch",
    "serial no lote": "batch_serial",
    "batch_serial": "batch_serial",
    "serial": "batch_serial",
    "fabricante": "manufacturer",
    "manufacturer": "manufacturer",
}


def import_cards(file_content: bytes) -> dict:
    from apps.cards.models import Card

    rows, errors = parse_excel_upload(file_content, ["card_uid"], header_map=CARD_HEADER_MAP)
    imported = 0

    for i, row in enumerate(rows, start=2):
        uid = row["card_uid"]
        if Card.objects.filter(card_uid=uid).exists():
            errors.append({"row": i, "detail": f"Cartao {uid} ja existe."})
            continue
        Card.objects.create(
            card_type=Card.CardType.PHYSICAL,
            card_uid=uid,
            card_number=row.get("card_number", ""),
            card_technology=row.get("card_technology", "nfc_uid"),
            issued_batch=row.get("issued_batch", ""),
            batch_serial=row.get("batch_serial", ""),
            manufacturer=row.get("manufacturer", ""),
            status=Card.Status.INACTIVE,
        )
        imported += 1

    return {"imported": imported, "errors": errors}


def import_stops(file_content: bytes) -> dict:
    from apps.routes.models import Stop

    rows, errors = parse_excel_upload(file_content, ["name"])
    imported = 0

    for i, row in enumerate(rows, start=2):
        Stop.objects.create(
            code=row.get("code", ""),
            name=row["name"],
            latitude=row.get("latitude") or None,
            longitude=row.get("longitude") or None,
        )
        imported += 1

    return {"imported": imported, "errors": errors}


def import_routes(file_content: bytes) -> dict:
    from apps.routes.models import Route

    rows, errors = parse_excel_upload(file_content, ["name"])
    imported = 0

    for i, row in enumerate(rows, start=2):
        Route.objects.create(
            code=row.get("code", ""),
            name=row["name"],
            description=row.get("description", ""),
        )
        imported += 1

    return {"imported": imported, "errors": errors}


def generate_card_template_excel() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cartoes NFC"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D3B66", end_color="0D3B66", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D4D4D8"),
        right=Side(style="thin", color="D4D4D8"),
        top=Side(style="thin", color="D4D4D8"),
        bottom=Side(style="thin", color="D4D4D8"),
    )

    headers = [
        ("card_uid", "UID do Cartao (obrigatorio)"),
        ("issued_batch", "Lote"),
        ("batch_serial", "Serial no Lote"),
        ("manufacturer", "Fabricante"),
    ]
    for col, (key, label) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    examples = [
        ["A1B2C3D4E5F6", "LOTE-2026-001", "001", "MIFARE"],
        ["G7H8I9J0K1L2", "LOTE-2026-001", "002", "MIFARE"],
        ["M3N4O5P6Q7R8", "LOTE-2026-001", "003", "NXP"],
    ]
    for r, row_data in enumerate(examples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _importar_vendas(file_content: bytes) -> dict:
    from apps.core.import_vendas import import_vendas_historicas

    return import_vendas_historicas(file_content)


IMPORTERS = {
    "cards": import_cards,
    "stops": import_stops,
    "routes": import_routes,
    "sales": _importar_vendas,
}


def generate_sales_template_excel() -> bytes:
    """Modelo para o operador preencher com as vendas ja realizadas.

    O ficheiro leva as colunas certas, um exemplo preenchido e uma folha a
    explicar o que cada coluna quer dizer — porque um modelo sem instrucoes
    volta preenchido a maneira de quem o recebeu, e cada linha mal preenchida
    e uma venda que nao entra nos relatorios.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D3B66", end_color="0D3B66", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    lado = Side(style="thin", color="D4D4D8")
    thin_border = Border(left=lado, right=lado, top=lado, bottom=lado)

    colunas = [
        ("Referencia", 22), ("Data", 14), ("Rota", 26), ("Origem", 22),
        ("Destino", 22), ("Passageiro", 26), ("Documento", 18),
        ("Telefone", 16), ("Valor", 12), ("Metodo", 14), ("Lugar", 8),
    ]
    for i, (titulo, largura) in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=i, value=titulo)
        celula.font = header_font
        celula.fill = header_fill
        celula.alignment = header_align
        celula.border = thin_border
        ws.column_dimensions[celula.column_letter].width = largura

    exemplo = ["TPM-2025-000123", "2025-11-14", "RT-MAPUTO-X-NELSPRUIT",
               "Polana Shopping", "Ilanga Mall", "Antonio Joaquim",
               "110100234567B", "841234567", "1650.00", "Dinheiro", "12A"]
    for i, valor in enumerate(exemplo, start=1):
        ws.cell(row=2, column=i, value=valor).border = thin_border
    ws.freeze_panes = "A2"

    guia = wb.create_sheet("Como preencher")
    guia.column_dimensions["A"].width = 18
    guia.column_dimensions["B"].width = 92
    linhas = [
        ("Coluna", "O que escrever"),
        ("Referencia", "OBRIGATORIO. O numero do bilhete no vosso sistema antigo. "
                       "E por ele que se sabe que uma linha ja foi carregada — carregar "
                       "o mesmo ficheiro duas vezes nao duplica nada."),
        ("Data", "OBRIGATORIO. O dia da viagem. Aceita 2025-11-14 ou 14/11/2025."),
        ("Rota", "O codigo ou o nome da rota, tal como esta na plataforma. Se nao "
                 "existir, a linha e recusada com o motivo — nao se inventa a rota."),
        ("Origem / Destino", "Nomes das paragens. Ficam escritos no bilhete historico."),
        ("Passageiro", "Nome de quem viajou."),
        ("Documento", "BI, passaporte ou DIRE. Opcional."),
        ("Telefone", "Numero de quem comprou. Opcional, mas e por ele que se "
                     "encontra o passageiro depois."),
        ("Valor", "OBRIGATORIO. Em meticais. Aceita 1650.00 ou 1.650,00."),
        ("Metodo", "Dinheiro, M-Pesa, e-Mola, Cartao ou Transferencia."),
        ("Lugar", "Numero do assento. Opcional."),
        ("", ""),
        ("Importante", "Os bilhetes carregados nascem JA USADOS: sao viagens que ja "
                       "aconteceram. Nao servem para viajar e nao e enviado nenhum SMS "
                       "a ninguem."),
    ]
    for i, (a, b) in enumerate(linhas, start=1):
        ca = guia.cell(row=i, column=1, value=a)
        cb = guia.cell(row=i, column=2, value=b)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 1 or a == "Importante":
            ca.font = Font(bold=True)
            cb.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
