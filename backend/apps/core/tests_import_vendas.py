"""Carregar para a plataforma as vendas que ja aconteceram.

O operador traz anos de bilhetes vendidos noutro sistema. O historico so serve
para alguma coisa se os relatorios fecharem — e so e seguro se nenhum desses
bilhetes puder ser usado para viajar.
"""

from __future__ import annotations

import io
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from apps.core.import_vendas import import_vendas_historicas
from apps.guest_checkouts.models import DigitalTravelPass, GuestCheckout
from apps.payments.models import PaymentIntent
from apps.routes.models import Route

CABECALHO = ["Referencia", "Data", "Rota", "Origem", "Destino", "Passageiro",
             "Documento", "Telefone", "Valor", "Metodo", "Lugar"]


def ficheiro(linhas) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(CABECALHO)
    for linha in linhas:
        ws.append(linha)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ImportarVendasTests(TestCase):
    def setUp(self):
        self.rota = Route.objects.create(
            code="RT-INT", name="Maputo x Nelspruit",
            service_type=Route.ServiceType.INTERNATIONAL, status=Route.Status.ACTIVE)

    def linha(self, ref="TPM-000123", data="2025-11-14", rota="RT-INT", valor="1650.00", **extra):
        return [ref, data, rota, extra.get("origem", "Polana"), extra.get("destino", "Ilanga"),
                extra.get("nome", "Antonio Joaquim"), extra.get("doc", "110100234567B"),
                extra.get("tel", "841234567"), valor, extra.get("metodo", "Dinheiro"),
                extra.get("lugar", "12A")]

    def test_uma_linha_vira_uma_venda_completa(self):
        r = import_vendas_historicas(ficheiro([self.linha()]))
        self.assertEqual(r["imported"], 1, r["errors"])
        self.assertEqual(r["errors"], [])
        self.assertEqual(r["total_amount"], "1650.00")

        checkout = GuestCheckout.objects.get()
        self.assertEqual(checkout.reference, "HIST-TPM-000123")
        self.assertEqual(checkout.total_amount, Decimal("1650.00"))
        self.assertEqual(checkout.buyer_name, "Antonio Joaquim")
        self.assertEqual(checkout.route_code, "RT-INT")

        bilhete = DigitalTravelPass.objects.get()
        self.assertEqual(bilhete.seat_number, "12A")
        self.assertEqual(bilhete.document_number, "110100234567B")

        pagamento = PaymentIntent.objects.get()
        self.assertEqual(pagamento.status, PaymentIntent.Status.CONFIRMED)
        self.assertEqual(pagamento.amount, Decimal("1650.00"))
        self.assertEqual(pagamento.provider, "cash")
        self.assertEqual(pagamento.channel, "import")

    def test_o_bilhete_historico_nao_serve_para_viajar(self):
        """Se nascesse activo, cada linha era um bilhete gratis a bordo."""
        import_vendas_historicas(ficheiro([self.linha()]))
        bilhete = DigitalTravelPass.objects.get()
        self.assertEqual(bilhete.status, DigitalTravelPass.Status.USED)
        self.assertIsNotNone(bilhete.used_at)
        self.assertLess(bilhete.valid_until, timezone.now())

    def test_a_venda_entra_nos_relatorios_na_data_em_que_aconteceu(self):
        """Nao na data em que se carregou o ficheiro."""
        import_vendas_historicas(ficheiro([self.linha(data="2025-11-14")]))
        pagamento = PaymentIntent.objects.get()
        self.assertEqual(timezone.localdate(pagamento.created_at).isoformat(), "2025-11-14")
        self.assertEqual(
            timezone.localdate(GuestCheckout.objects.get().created_at).isoformat(), "2025-11-14")

    def test_carregar_o_mesmo_ficheiro_duas_vezes_nao_duplica(self):
        """Senao a receita dobrava nos relatorios e ninguem reparava."""
        conteudo = ficheiro([self.linha(), self.linha(ref="TPM-000124")])
        primeira = import_vendas_historicas(conteudo)
        segunda = import_vendas_historicas(conteudo)

        self.assertEqual(primeira["imported"], 2)
        self.assertEqual(segunda["imported"], 0)
        self.assertEqual(segunda["duplicates"], 2)
        self.assertEqual(GuestCheckout.objects.count(), 2)
        self.assertEqual(PaymentIntent.objects.count(), 2)

    def test_nao_se_escreve_nem_se_cobra_a_ninguem(self):
        from apps.sms.models import SmsMessage

        import_vendas_historicas(ficheiro([self.linha()]))
        self.assertEqual(SmsMessage.objects.count(), 0)

    # --- o que as pessoas escrevem de facto -----------------------------

    def test_datas_em_varios_formatos(self):
        for data in ["2025-11-14", "14/11/2025", "14-11-2025", "2025-11-14 00:00:00"]:
            with self.subTest(data=data):
                GuestCheckout.all_objects.all().delete()
                DigitalTravelPass.all_objects.all().delete()
                PaymentIntent.all_objects.all().delete()
                r = import_vendas_historicas(ficheiro([self.linha(ref=f"R-{data}", data=data)]))
                self.assertEqual(r["imported"], 1, f"{data}: {r['errors']}")

    def test_valores_a_portuguesa_e_a_inglesa(self):
        for valor, esperado in [("1650.00", "1650.00"), ("1.650,00", "1650.00"),
                                ("1650", "1650.00"), ("1650,50", "1650.50")]:
            with self.subTest(valor=valor):
                GuestCheckout.all_objects.all().delete()
                DigitalTravelPass.all_objects.all().delete()
                PaymentIntent.all_objects.all().delete()
                import_vendas_historicas(ficheiro([self.linha(ref=f"V-{valor}", valor=valor)]))
                self.assertEqual(GuestCheckout.objects.get().total_amount, Decimal(esperado))

    def test_rota_pelo_nome_tambem_serve(self):
        r = import_vendas_historicas(ficheiro([self.linha(rota="Maputo x Nelspruit")]))
        self.assertEqual(r["imported"], 1, r["errors"])
        self.assertEqual(GuestCheckout.objects.get().route_code, "RT-INT")

    # --- o que tem de falhar, e dizer porque ----------------------------

    def test_rota_inexistente_para_a_linha_e_explica(self):
        r = import_vendas_historicas(ficheiro([self.linha(rota="RT-QUE-NAO-EXISTE")]))
        self.assertEqual(r["imported"], 0)
        self.assertIn("nao existe", r["errors"][0]["detail"])
        self.assertEqual(r["errors"][0]["row"], 2)

    def test_data_impossivel_para_a_linha(self):
        r = import_vendas_historicas(ficheiro([self.linha(data="14 de Novembro")]))
        self.assertEqual(r["imported"], 0)
        self.assertIn("Data invalida", r["errors"][0]["detail"])

    def test_valor_zero_ou_negativo_e_recusado(self):
        r = import_vendas_historicas(ficheiro([
            self.linha(ref="A", valor="0"), self.linha(ref="B", valor="-100")]))
        self.assertEqual(r["imported"], 0)
        self.assertEqual(len(r["errors"]), 2)

    def test_uma_linha_ma_nao_trava_as_boas(self):
        r = import_vendas_historicas(ficheiro([
            self.linha(ref="BOA-1"),
            self.linha(ref="MA", data="ontem"),
            self.linha(ref="BOA-2"),
        ]))
        self.assertEqual(r["imported"], 2)
        self.assertEqual(len(r["errors"]), 1)


class EndpointDeImportacaoTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.admin = User.objects.create_superuser(username="raiz", password="x", email="r@x.mz")
        self.client = APIClient()
        self.client.force_authenticate(self.admin)
        Route.objects.create(code="RT-INT", name="Rota", status=Route.Status.ACTIVE)

    def test_upload_pelo_portal(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        conteudo = ficheiro([["TPM-1", "2025-11-14", "RT-INT", "A", "B", "Ze",
                              "", "841000000", "500.00", "Dinheiro", ""]])
        ficheiro_enviado = SimpleUploadedFile(
            "vendas.xlsx", conteudo,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        r = self.client.post("/api/import/sales/", {"file": ficheiro_enviado}, format="multipart")
        self.assertEqual(r.status_code, 201, r.content)
        self.assertEqual(r.json()["imported"], 1)

    def test_modelo_para_o_cliente_preencher(self):
        r = self.client.get("/api/import/sales/template/")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertEqual([c.value for c in wb["Vendas"][1]], CABECALHO)
        self.assertIn("Como preencher", wb.sheetnames)

    def test_sem_permissao_nao_importa(self):
        User = get_user_model()
        ze = User.objects.create_user(username="ze", password="x", email="z@x.mz")
        self.client.force_authenticate(ze)
        self.assertEqual(self.client.get("/api/import/sales/template/").status_code, 403)
