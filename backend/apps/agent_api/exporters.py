"""PDF + Excel exporters for the agent revenue admin module.

Branded with BuzUp and powered-by UpDigital. Logos are
loaded from `backend/static/assets/` and embedded into the documents.
"""
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

NAVY = colors.HexColor("#071E49")
ORANGE = colors.HexColor("#E47B11")
GREY = colors.HexColor("#6B6356")
LIGHT_GREY = colors.HexColor("#E7E1D4")
SOFT_BG = colors.HexColor("#F7F4EE")


def _asset(*parts: str) -> Path:
    return Path(settings.BASE_DIR) / "static" / "assets" / Path(*parts)


def _safe_image(path: Path) -> ImageReader | None:
    try:
        if path.exists():
            return ImageReader(str(path))
    except Exception:
        return None
    return None


def _to_decimal(v) -> Decimal:
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0.00")


def _fmt_money(v) -> str:
    return f"{_to_decimal(v):,.2f} MZN"


# ---------------------------------------------------------------------------
# PDF: single day-close session
# ---------------------------------------------------------------------------

def session_pdf(record) -> bytes:
    """Generate a PDF for a single AgentDayClose.

    Layout:
      - Header band w/ BuzUp logo, title, date
      - Agent + period info
      - 4 KPI boxes (sales/topups/validations/totals)
      - Table for sales, topups, validations
      - Footer: powered by UpDigital
    """
    buf = io.BytesIO()
    page_size = A4
    width, height = page_size
    c = canvas.Canvas(buf, pagesize=page_size)

    payload = record.payload or {}
    totals = payload.get("totals", {})
    sales = payload.get("sales", [])
    topups = payload.get("topups", [])
    validations = payload.get("validations", [])

    _draw_header(c, width, height, title=f"Fecho de Caixa - {record.date.isoformat()}")
    y = height - 38 * mm

    # Agent line
    c.setFillColor(NAVY)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(15 * mm, y, f"Agente: {record.agent_user.get_full_name() or record.agent_user.username}")
    y -= 6 * mm
    c.setFont("Helvetica", 10)
    c.setFillColor(GREY)
    closed_at = record.closed_at.strftime("%Y-%m-%d %H:%M")
    c.drawString(15 * mm, y, f"Fechado em: {closed_at}    Sessoes: {record.sessions_closed}")
    y -= 10 * mm

    # KPI boxes (4 across)
    box_w = (width - 30 * mm - 9) / 4
    box_h = 22 * mm
    kpis = [
        ("VENDAS", _fmt_money(record.sales_total), f"{record.tickets_count} bilhetes", ORANGE),
        ("RECARGAS", _fmt_money(record.topups_total), f"{len(topups)} operacoes", colors.HexColor("#0B6FE0")),
        ("VALIDACOES", str(record.validations_count), _fmt_money(record.validations_revenue), colors.HexColor("#8B5CF6")),
        ("RECEITA EM CAIXA", _fmt_money(_to_decimal(record.sales_total) + _to_decimal(record.topups_total)), "Vendas + Recargas", colors.HexColor("#1FB04A")),
    ]
    x = 15 * mm
    for label, value, foot, accent in kpis:
        c.setStrokeColor(LIGHT_GREY)
        c.setFillColor(colors.white)
        c.roundRect(x, y - box_h, box_w, box_h, 4, fill=1, stroke=1)
        c.setFillColor(accent)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(x + 4, y - 6, label)
        c.setFillColor(NAVY)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x + 4, y - 14, value)
        c.setFillColor(GREY)
        c.setFont("Helvetica", 8)
        c.drawString(x + 4, y - box_h + 4, foot)
        x += box_w + 3

    y -= box_h + 8 * mm

    # ---- Sales table ----
    y = _draw_table(
        c, x_left=15 * mm, y=y, width=width - 30 * mm,
        title=f"Vendas ({len(sales)})",
        headers=["Referencia", "Telefone", "Qtd", "Valor", "Estado"],
        col_widths=[60, 35, 15, 25, 25],
        rows=[[
            str(r.get("sale_reference") or r.get("reference") or "-"),
            str(r.get("payer_phone_masked") or "-"),
            str(r.get("quantity") or "-"),
            _fmt_money(r.get("amount", 0)),
            str(r.get("status") or "-").upper(),
        ] for r in sales[:30]],
        empty_text="Sem vendas neste fecho.",
    )

    # ---- Topups table ----
    y = _draw_table(
        c, x_left=15 * mm, y=y, width=width - 30 * mm,
        title=f"Recargas ({len(topups)})",
        headers=["Referencia", "Telefone", "Valor", "Estado"],
        col_widths=[60, 35, 30, 35],
        rows=[[
            str(r.get("reference") or "-"),
            str(r.get("payer_phone_masked") or "-"),
            _fmt_money(r.get("amount", 0)),
            str(r.get("status") or "-").upper(),
        ] for r in topups[:30]],
        empty_text="Sem recargas neste fecho.",
    )

    # ---- Validations table ----
    y = _draw_table(
        c, x_left=15 * mm, y=y, width=width - 30 * mm,
        title=f"Validacoes ({len(validations)})",
        headers=["Tipo", "Rota", "Debito", "Dispositivo", "Estado"],
        col_widths=[35, 25, 28, 42, 30],
        rows=[[
            str(r.get("validation_type") or "-")[:20],
            str(r.get("route") or "-")[:14],
            _fmt_money(r.get("amount_debited", 0)),
            str(r.get("device_serial") or "-")[:24],
            str(r.get("status") or "-").upper(),
        ] for r in validations[:30]],
        empty_text="Sem validacoes neste fecho.",
    )

    _draw_footer(c, width)
    c.save()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF: aggregated summary
# ---------------------------------------------------------------------------

def summary_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    page_size = landscape(A4)
    width, height = page_size
    c = canvas.Canvas(buf, pagesize=page_size)

    totals = data.get("totals", {})
    agents = data.get("agents", [])
    df = data.get("date_from", "")
    dt = data.get("date_to", "")

    _draw_header(c, width, height, title=f"Resumo de Receita por Agente | {df} a {dt}")
    y = height - 36 * mm

    # KPIs
    box_w = (width - 30 * mm - 12) / 5
    box_h = 22 * mm
    kpis = [
        ("RECEITA TOTAL", _fmt_money(totals.get("total_revenue", 0)), "Vendas + Recargas", ORANGE),
        ("VENDAS", _fmt_money(totals.get("sales_total", 0)), f"{totals.get('tickets', 0)} bilhetes", colors.HexColor("#1FB04A")),
        ("RECARGAS", _fmt_money(totals.get("topups_total", 0)), "Carteiras carregadas", colors.HexColor("#0B6FE0")),
        ("VALIDACOES", str(totals.get("validations", 0)), _fmt_money(totals.get("validations_revenue", 0)), colors.HexColor("#8B5CF6")),
        ("AGENTES", str(totals.get("agents_count", 0)), f"{totals.get('closes', 0)} fechos", colors.HexColor("#EAB308")),
    ]
    x = 15 * mm
    for label, value, foot, accent in kpis:
        c.setStrokeColor(LIGHT_GREY); c.setFillColor(colors.white)
        c.roundRect(x, y - box_h, box_w, box_h, 4, fill=1, stroke=1)
        c.setFillColor(accent); c.setFont("Helvetica-Bold", 8); c.drawString(x + 4, y - 6, label)
        c.setFillColor(NAVY); c.setFont("Helvetica-Bold", 14); c.drawString(x + 4, y - 14, value)
        c.setFillColor(GREY); c.setFont("Helvetica", 8); c.drawString(x + 4, y - box_h + 4, foot)
        x += box_w + 3

    y -= box_h + 8 * mm

    # Per-agent table
    y = _draw_table(
        c, x_left=15 * mm, y=y, width=width - 30 * mm,
        title="Por agente",
        headers=["Agente", "Telefone", "Vendas", "Recargas", "Receita", "Validacoes (#)", "Bilhetes", "Fechos"],
        col_widths=[55, 30, 28, 28, 30, 30, 22, 18],
        rows=[[
            str(a.get("agent_name") or "-")[:30],
            str(a.get("agent_phone") or "-"),
            _fmt_money(a.get("sales_total", 0)),
            _fmt_money(a.get("topups_total", 0)),
            _fmt_money(a.get("total_revenue", 0)),
            str(a.get("validations", 0)),
            str(a.get("tickets", 0)),
            str(a.get("closes", 0)),
        ] for a in agents[:30]],
        empty_text="Sem actividade no periodo.",
    )

    _draw_footer(c, width)
    c.save()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel exports
# ---------------------------------------------------------------------------

def session_xlsx(record) -> bytes:
    wb = Workbook()
    payload = record.payload or {}

    # Resumo sheet
    s = wb.active
    s.title = "Resumo"
    _xlsx_set_header(s, f"Fecho de caixa - {record.date.isoformat()}")
    s["A3"] = "Agente"
    s["B3"] = record.agent_user.get_full_name() or record.agent_user.username
    s["A4"] = "Fechado em"
    s["B4"] = record.closed_at.strftime("%Y-%m-%d %H:%M")
    s["A5"] = "Sessoes encerradas"
    s["B5"] = record.sessions_closed

    s["A7"] = "Receita em caixa (Vendas + Recargas)"
    s["B7"] = float(_to_decimal(record.sales_total) + _to_decimal(record.topups_total))
    s["A8"] = "Vendas"
    s["B8"] = float(record.sales_total)
    s["A9"] = "Recargas"
    s["B9"] = float(record.topups_total)
    s["A10"] = "Bilhetes emitidos"
    s["B10"] = record.tickets_count
    s["A11"] = "Validacoes (numero)"
    s["B11"] = record.validations_count
    s["A12"] = "Validacoes (valor debitado em carteiras)"
    s["B12"] = float(record.validations_revenue)
    for r in range(7, 13):
        s[f"A{r}"].font = Font(bold=True)
    s.column_dimensions["A"].width = 42
    s.column_dimensions["B"].width = 22

    # Vendas sheet
    sales = payload.get("sales", [])
    _xlsx_table(wb, "Vendas", ["Referencia", "Telefone", "Quantidade", "Valor", "Estado", "Criado em"],
                [[
                    r.get("sale_reference") or r.get("reference") or "",
                    r.get("payer_phone_masked") or "",
                    r.get("quantity") or 0,
                    float(_to_decimal(r.get("amount", 0))),
                    str(r.get("status") or "").upper(),
                    r.get("created_at") or "",
                ] for r in sales])

    # Recargas sheet
    topups = payload.get("topups", [])
    _xlsx_table(wb, "Recargas", ["Referencia", "Telefone", "Valor", "Estado", "Criado em"],
                [[
                    r.get("reference") or "",
                    r.get("payer_phone_masked") or "",
                    float(_to_decimal(r.get("amount", 0))),
                    str(r.get("status") or "").upper(),
                    r.get("created_at") or "",
                ] for r in topups])

    # Validacoes sheet
    validations = payload.get("validations", [])
    _xlsx_table(wb, "Validacoes", ["Tipo", "Rota", "Debito (MZN)", "Dispositivo", "Estado", "Criado em"],
                [[
                    r.get("validation_type") or "",
                    r.get("route") or "",
                    float(_to_decimal(r.get("amount_debited", 0))),
                    r.get("device_serial") or "",
                    str(r.get("status") or "").upper(),
                    r.get("created_at") or "",
                ] for r in validations])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def summary_xlsx(data: dict) -> bytes:
    wb = Workbook()
    totals = data.get("totals", {})
    agents = data.get("agents", [])

    s = wb.active
    s.title = "Totais"
    _xlsx_set_header(s, f"Receita por agente | {data.get('date_from', '')} a {data.get('date_to', '')}")
    rows = [
        ("Receita total (vendas + recargas)", float(_to_decimal(totals.get("total_revenue", 0)))),
        ("Vendas", float(_to_decimal(totals.get("sales_total", 0)))),
        ("Recargas", float(_to_decimal(totals.get("topups_total", 0)))),
        ("Validacoes (numero)", int(totals.get("validations", 0) or 0)),
        ("Validacoes (valor debitado)", float(_to_decimal(totals.get("validations_revenue", 0)))),
        ("Bilhetes emitidos", int(totals.get("tickets", 0) or 0)),
        ("Agentes activos", int(totals.get("agents_count", 0) or 0)),
        ("Fechos submetidos", int(totals.get("closes", 0) or 0)),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        s.cell(row=i, column=1, value=label).font = Font(bold=True)
        s.cell(row=i, column=2, value=value)
    s.column_dimensions["A"].width = 44
    s.column_dimensions["B"].width = 24

    _xlsx_table(wb, "Por agente",
                ["Agente", "Telefone", "Vendas", "Recargas", "Receita", "Validacoes (#)", "Validacoes (MZN)", "Bilhetes", "Fechos"],
                [[
                    a.get("agent_name") or "",
                    a.get("agent_phone") or "",
                    float(_to_decimal(a.get("sales_total", 0))),
                    float(_to_decimal(a.get("topups_total", 0))),
                    float(_to_decimal(a.get("total_revenue", 0))),
                    int(a.get("validations", 0) or 0),
                    float(_to_decimal(a.get("validations_revenue", 0))),
                    int(a.get("tickets", 0) or 0),
                    int(a.get("closes", 0) or 0),
                ] for a in agents])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Drawing primitives
# ---------------------------------------------------------------------------

def _draw_header(c: canvas.Canvas, width: float, height: float, *, title: str) -> None:
    """Top band: BuzUp mark + title + 'BuzUp - Plataforma cashless'."""
    band_h = 22 * mm
    c.setFillColor(NAVY)
    c.rect(0, height - band_h, width, band_h, fill=1, stroke=0)

    # Logo (left)
    logo = _safe_image(_asset("buzup-logo", "buzup_dark.png")) or _safe_image(_asset("buzup-logo", "buzup_light.png"))
    if logo:
        try:
            iw, ih = logo.getSize()
            target_h = 14 * mm
            target_w = iw * target_h / ih
            c.drawImage(logo, 8 * mm, height - band_h + (band_h - target_h) / 2, width=target_w, height=target_h, mask="auto")
        except Exception:
            pass

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(60 * mm, height - 10 * mm, "BuzUp - Receita do Agente")
    c.setFont("Helvetica", 9)
    c.drawString(60 * mm, height - 15 * mm, "BuzUp | Transporte cashless de Mocambique")

    c.setFillColor(ORANGE)
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(width - 10 * mm, height - 10 * mm, title)
    c.setFillColor(colors.white)
    c.setFont("Helvetica", 8)
    c.drawRightString(width - 10 * mm, height - 15 * mm, f"Gerado em {datetime.now().strftime('%Y-%m-%d %H:%M')}")


def _draw_footer(c: canvas.Canvas, width: float) -> None:
    band_h = 14 * mm
    c.setFillColor(SOFT_BG)
    c.rect(0, 0, width, band_h, fill=1, stroke=0)
    up = _safe_image(_asset("up-digital-logo", "up_digital_dark.png")) or _safe_image(_asset("up-digital-logo", "up_digital_light.png"))
    if up:
        try:
            iw, ih = up.getSize()
            target_h = 7 * mm
            target_w = iw * target_h / ih
            c.drawImage(up, width - 8 * mm - target_w, (band_h - target_h) / 2, width=target_w, height=target_h, mask="auto")
        except Exception:
            pass
    c.setFillColor(GREY)
    c.setFont("Helvetica", 8)
    c.drawString(10 * mm, band_h / 2 - 2, "BuzUp | Documento gerado automaticamente.")
    c.drawString(10 * mm, band_h / 2 - 9, "powered by")


def _draw_table(c, *, x_left, y, width, title, headers, col_widths, rows, empty_text):
    line_h = 5.5 * mm
    if y - (len(rows) + 4) * line_h < 24 * mm:
        c.showPage()
        # Header on next page
        page_size = c._pagesize
        page_w, page_h = page_size
        _draw_header(c, page_w, page_h, title="Continuacao")
        y = page_h - 30 * mm

    c.setFillColor(NAVY)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x_left, y, title)
    y -= 6
    # Header row
    cw_sum = sum(col_widths)
    scale = width / cw_sum
    x = x_left
    c.setFillColor(NAVY)
    c.rect(x_left, y - line_h, width, line_h, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    for h, w in zip(headers, col_widths):
        c.drawString(x + 2, y - line_h + 1.5 * mm, h)
        x += w * scale
    y -= line_h

    if not rows:
        c.setFillColor(GREY)
        c.setFont("Helvetica-Oblique", 8)
        c.drawString(x_left + 4, y - 4 * mm, empty_text)
        return y - 8 * mm

    c.setFont("Helvetica", 8)
    for i, row in enumerate(rows):
        if i % 2 == 1:
            c.setFillColor(SOFT_BG)
            c.rect(x_left, y - line_h, width, line_h, fill=1, stroke=0)
        x = x_left
        c.setFillColor(NAVY)
        for cell, w in zip(row, col_widths):
            txt = str(cell)
            c.drawString(x + 2, y - line_h + 1.5 * mm, txt)
            x += w * scale
        y -= line_h
    return y - 4 * mm


def _xlsx_set_header(ws, title: str):
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True, color="071E49")
    ws.merge_cells("A1:F1")


def _xlsx_table(wb, sheet_name: str, headers: list, rows: list):
    ws = wb.create_sheet(title=sheet_name)
    head_fill = PatternFill("solid", fgColor="071E49")
    head_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style="thin", color="DDDDDD"),
                    right=Side(style="thin", color="DDDDDD"),
                    top=Side(style="thin", color="DDDDDD"),
                    bottom=Side(style="thin", color="DDDDDD"))
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_i)].width = max(14, min(40, len(h) + 8))
    for r_i, r in enumerate(rows, start=2):
        for c_i, v in enumerate(r, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=v)
            cell.border = border
            if isinstance(v, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                if c_i >= 3:
                    cell.number_format = "#,##0.00"
    ws.freeze_panes = "A2"
